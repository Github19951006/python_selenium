"""
@Project ：python 
@Author : 文跃锐（yuerwen）
@Time   : 2021/11/12
@File   :yuerwen_openpyxl_背景颜色.py
"""
import openpyxl
from openpyxl.styles import PatternFill

# 加载Excel的book对象
book = openpyxl.load_workbook('income.xlsx')

# 通过sheet的名字获取 sheet对象
sheet = book['2018']

# 更改指定单元格的背景颜色  rgb的背景颜色
sheet['A1'].fill = PatternFill('solid','E39191')

#更改整行或者整列的背景颜色  只能使用循环
fill = PatternFill('solid','E39191')
for one in range(1,20):
	# 指定固定的列 ，更改行的位置索引
	sheet.cell(row = one,column= 1).fill = fill
	
# 保存文件
book.save('income.xlsx')