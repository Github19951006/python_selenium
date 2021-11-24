"""
@Project ：python
@Author : 文跃锐（yuerwen）
@Time   : 2021/11/08
@File   :yuerwen_Excel_写入数据1.py
"""
import openpyxl

# 创建一个Excel workbook对象
book = openpyxl.Workbook()

# 创建时，会创建一个sheet对象
sheet = book.active

# 给sheet表取一个表名
sheet.title = '工作表'

book.save('新创建.xlsx')

# 增加一个新的sheet表，放在最后
sh1 = book.create_sheet('年龄表-放在最后')

# 增加一个新的sheet表，放在最前
sh2 = book.create_sheet('开始表-放在最前',0)

# 增加一个sheet表，放在指定的位置
sh3 = book.create_sheet('指定位置的表-第二个位置',1)

# 根据名称获取某个sheet对象
sheet = book['工作表']

# 给第一个单元格写入内容
sheet['A1'] = '你好'

# 获取某个单元格内容
print(sheet['A1'].value)

# 根据行号列号， 给第一个单元格写入内容，
# 注意和 xlrd 不同，是从 1 开始
sheet.cell(2,2).value = '白月黑羽'

# 根据行号列号， 获取某个单元格内容
print(sheet.cell(2, 2).value)
# 存储Excel表文件
book.save('新创建.xlsx')