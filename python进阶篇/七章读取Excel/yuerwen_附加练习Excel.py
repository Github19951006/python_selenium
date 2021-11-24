"""
@Project ：python 
@Author : 文跃锐（yuerwen）
@Time   : 2021/11/17
@File   :yuerwen_附加练习Excel.py
"""

import json
import time
import openpyxl


def open_write_Excel():
	# 数据字段映射表
	fieldMap = {
		'dynamicTime': '计划时间',
		'shipNameCn': '中文船名',
		'shipNameEn': '英文船名',
		'shipFlag': '国籍(地区)',
		'shipLength': '船长',
		'draft': '吃水',
		'dynamicName': '动态',
		'startBerth': '起点泊位',
		'endBerth': '终点泊位',
		'master': '主引',
		'assistant': '副引',
		'assistant2': '其它引水',
		'orgShort': '代理',
		'dTelephone': '代理电话',
		'+++1': '交通船',
		'channelName': '航道',
		'headThruster:tailThruster': '侧推',
		'remarks': '备注',
	}
	
	# 创建Excel的book对象
	book = openpyxl.Workbook()
	# 在创建时，自动生成的sheet对象 ，通话active获取
	sheet = book.active
	# 对当前sheet表的命名
	sheet.title = '航班表'
	
	# 写入标题栏
	for nub, val in enumerate(fieldMap.values()):
		sheet.cell(1, nub + 1).value = val
	
	# 读文件
	with open('info1.txt', 'r', encoding='utf8') as f:
		row = 1
		while True:
			# 读入一行
			oneline = f.readline()
			# 判空
			if not oneline:
				break
			# # 一行对应1天的数据
			oneday = json.loads(oneline)
			
			# 再获取当天 的每行数据
			for line in oneday:
				row += 1
				# 过滤符合的日期
				if int(line['dynamicTime'] / 1000) <= 1546934400:
					for column, field in enumerate(fieldMap.keys()):
						# 没有的，暂时填空
						if field not in line:
							value = ''
						else:
							# 通过key拿value的值
							value = line[field]
						
						# 计划时间要特殊处理
						if field == 'dynamicTime':
							# 将秒格式化成自定义模式
							value = time.strftime('%Y%m%d %H:%M', time.localtime(int(line['dynamicTime'] / 1000)))
						
						# 侧推要特殊处理
						if field == 'headThruster:tailThruster':
							if 'headThruster' not in line:
								hvalue = '无'
							else:
								hvalue = line['headThruster']
							
							if 'tailThruster' not in line:
								tvalue = '无'
							else:
								tvalue = line['tailThruster']
							
							value = f"首：{hvalue} 尾：{tvalue}"
						
						sheet.cell(row, column + 1).value = value
		# 保存文件
		book.save('引航表_附加.xls')


open_write_Excel()
