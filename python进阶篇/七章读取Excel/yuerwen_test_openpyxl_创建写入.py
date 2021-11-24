"""
@Project ：python 
@Author : 文跃锐（yuerwen）
@Time   : 2021/11/10
@File   :yuerwen_test_openpyxl_创建写入.py
"""
import openpyxl

name2Age = {
    '张飞' :  38,
    '赵云' :  27,
    '许褚' :  36,
    '典韦' :  38,
    '关羽' :  39,
    '黄忠' :  49,
    '徐晃' :  43,
    '马超' :  23,
}

# 创建一个book对象
book = openpyxl.Workbook()

# 创建时，生成一个sheet对象
sheet = book.active

# 给表取一个名字
sheet.title = 'name_age表'

# 写表题栏
sheet['A1'] = 'name'
sheet['B1'] = 'age'

# 创建一个行的迭代器
row = 2

# 写入内容
for name,age in name2Age.items():
	sheet.cell(row,1).value = name
	sheet.cell(row,2).value = age
	row += 1
	
# 保存book文件
book.save('员工信息.xlsx')


